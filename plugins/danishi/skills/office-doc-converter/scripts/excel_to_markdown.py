#!/usr/bin/env python3
"""Convert Excel workbooks to structured Markdown with sheet separation and image links."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def detect_grid_style(ws: Worksheet) -> bool:
    """Heuristic to detect Excel方眼紙 (grid-style layout)."""
    col_count = ws.max_column
    row_count = ws.max_row
    has_many_columns = col_count >= 20
    squareish = 0.5 <= (col_count / max(row_count, 1)) <= 2
    return has_many_columns and squareish


def merged_cell_summary(ws: Worksheet) -> list[str]:
    return [str(rng) for rng in ws.merged_cells.ranges]


def extract_sheet_images(ws: Worksheet, target_root: Path) -> list[Path]:
    extracted: list[Path] = []
    if not hasattr(ws, "_images"):
        return extracted

    sheet_dir = target_root / ws.title
    sheet_dir.mkdir(parents=True, exist_ok=True)

    for idx, image in enumerate(getattr(ws, "_images", []), start=1):
        filename = f"{ws.title}-image{idx}.png"
        target_path = sheet_dir / filename
        try:
            target_path.write_bytes(image._data())  # type: ignore[attr-defined]
        except Exception as exc:  # noqa: BLE001
            print(f"Warning: failed to extract image {idx} on sheet {ws.title}: {exc}")
            continue
        extracted.append(target_path)
    return extracted


def dataframe_to_markdown(df: pd.DataFrame) -> str:
    normalized = df.fillna("")
    return normalized.to_markdown(index=False)


def build_sheet_markdown(sheet_name: str, df: pd.DataFrame, merged_ranges: list[str], grid_style: bool, images: list[Path]) -> str:
    lines = [f"## {sheet_name}"]
    lines.append(f"- サイズ: {df.shape[0]} 行 x {df.shape[1]} 列")
    lines.append(f"- Excel方眼紙の可能性: {'あり' if grid_style else 'なし'}")
    if merged_ranges:
        lines.append("- セル結合: " + ", ".join(merged_ranges))
    else:
        lines.append("- セル結合: なし")

    if images:
        lines.append("- 画像一覧:")
        for path in images:
            lines.append(f"  - ![{sheet_name}]({path.as_posix()})")

    lines.append("")
    lines.append(dataframe_to_markdown(df))
    lines.append("")
    return "\n".join(lines)


def convert_excel_to_markdown(input_path: Path, output_path: Path, media_dir: Path | None = None) -> None:
    workbook = load_workbook(input_path, data_only=True)
    sheet_frames = pd.read_excel(input_path, sheet_name=None, dtype=str)

    media_links: dict[str, list[Path]] = {}
    if media_dir is not None:
        for ws in workbook.worksheets:
            extracted = extract_sheet_images(ws, media_dir / input_path.stem)
            if extracted:
                media_links[ws.title] = extracted

    lines: list[str] = [f"# Excel変換結果: {input_path.name}", ""]
    lines.append(f"- シート数: {len(sheet_frames)}")
    lines.append(f"- 出力日時: auto-generated")
    lines.append("")

    for sheet_name, df in sheet_frames.items():
        ws = workbook[sheet_name]
        merged_ranges = merged_cell_summary(ws)
        grid_style = detect_grid_style(ws)
        images = media_links.get(sheet_name, [])
        lines.append(build_sheet_markdown(sheet_name, df, merged_ranges, grid_style, images))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert Excel workbooks to Markdown")
    parser.add_argument("input_excel", type=Path, help="Path to the Excel workbook")
    parser.add_argument("output_markdown", type=Path, help="Destination Markdown file")
    parser.add_argument("--media-dir", type=Path, help="Directory to save extracted images", default=None)
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)

    if not args.input_excel.exists():
        print(f"Input Excel not found: {args.input_excel}", file=sys.stderr)
        return 1

    convert_excel_to_markdown(args.input_excel, args.output_markdown, args.media_dir)
    print(f"Markdown written to {args.output_markdown}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
